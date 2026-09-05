#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BRING 주간 업무지시서 생성기 (Claude Code / Codex 공용)

빈 마스터 템플릿(BRING_주간업무지시서_템플릿(빈양식).xlsx)에 입력(JSON)을 채워
담당자별 주간 업무지시서 xlsx를 생성한다.

특징
- 오늘 날짜 기준 '현재 주차'로 주차·대상기간·보고기한·마감을 자동 최신화
  (input에서 week_monday를 주면 그 주로 고정, null 이면 이번 주)
- 업무 개수(N)에 맞춰 §3 지시업무 표와 02_주간완료보고서 시트를 자동 확장/축소
  (수식·병합·데이터검증·결재란 위치 재정합)
- 가중치 합계 100% 검증

사용법
  python generate_directive.py directive_input.example.json
  python generate_directive.py input.json -o 결과.xlsx --template "BRING_주간업무지시서_템플릿(빈양식).xlsx"
"""
import sys
import json
import argparse
import datetime
from copy import copy

import openpyxl
from openpyxl.utils.cell import range_boundaries, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

COLS = range(1, 14)  # A..M
WK = ["월", "화", "수", "목", "금", "토", "일"]
DEFAULT_TEMPLATE = "BRING_주간업무지시서_템플릿(빈양식).xlsx"


def week_info(week_monday=None):
    """현재(또는 지정) 주차의 날짜 정보를 계산한다."""
    if week_monday:
        d = datetime.date.fromisoformat(week_monday)
    else:
        d = datetime.date.today()
    mon = d - datetime.timedelta(days=d.weekday())   # 그 주 월요일
    fri = mon + datetime.timedelta(days=4)
    iso = mon.isocalendar()
    f = lambda x: x.strftime("%m-%d")
    return {
        "week_str": f"{iso[0]}-W{iso[1]:02d}",
        "year": iso[0], "week": iso[1],
        "period": f"{f(mon)}(월)~{f(fri)}(금)",
        "report_due": f"{f(fri)}(금) 17:00",
        "ack_due": f"{f(mon)}(월) 12:00",
        "friday": f"{f(fri)}(금) 17:00",
        "approve_date": mon.isoformat(),
    }


def _expand(ws, insert_before, M, tmpl_row, new_h):
    """insert_before 위치에 M개 행을 삽입하고 아래 내용을 밀어내며,
    병합·데이터검증·행높이·스타일을 재정합한다. (M<=0이면 아무것도 안 함)"""
    if M <= 0:
        return
    merges = [str(m) for m in ws.merged_cells.ranges]
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))
    ws.data_validations.dataValidation = []
    maxr = ws.max_row

    def move(s, d):
        for c in COLS:
            a = ws.cell(row=s, column=c)
            b = ws.cell(row=d, column=c)
            b.value = a.value
            if a.has_style:
                b._style = copy(a._style)
        h = ws.row_dimensions[s].height
        if h is not None:
            ws.row_dimensions[d].height = h

    for r in range(maxr, insert_before - 1, -1):
        move(r, r + M)
    for r in range(insert_before, insert_before + M):
        for c in COLS:
            ws.cell(row=r, column=c).value = None
            a = ws.cell(row=tmpl_row, column=c)
            if a.has_style:
                ws.cell(row=r, column=c)._style = copy(a._style)
        ws.row_dimensions[r].height = new_h

    def sh(ref):
        c1, r1, c2, r2 = range_boundaries(ref)
        if r1 >= insert_before:
            r1 += M
            r2 += M
        return f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"

    for m in merges:
        ws.merge_cells(sh(m))


def build(inp, out_path=None, template=None):
    info = week_info(inp.get("week_monday"))
    tasks = inp["tasks"]
    N = len(tasks)
    if N < 1:
        raise ValueError("tasks 는 최소 1개 이상이어야 합니다.")
    M = max(0, N - 5)
    last = 27 + N          # 마지막 지시업무 행
    lr = 11 + N            # 완료보고서 마지막 업무 행

    wb = openpyxl.load_workbook(template or inp.get("template", DEFAULT_TEMPLATE))
    w1 = wb["01_업무지시서"]
    w2 = wb["02_주간완료보고서"]

    _expand(w1, 33, M, 28, 90.0)
    _expand(w2, 17, M, 12, 27.75)

    # N<5 이면 남는 업무행을 비운다
    if N < 5:
        for r in range(28 + N, 33):
            for c in COLS:
                w1.cell(row=r, column=c).value = None
        for r in range(12 + N, 17):
            for c in COLS:
                w2.cell(row=r, column=c).value = None

    sum1, issue1, sign1 = 33 + M, 34 + M, 47 + M   # 합계 / 발행검증 / 결재 서명행
    sum2 = 17 + M                                    # 완료보고서 합계행

    def S(ws, coord, val, pct=False):
        ws[coord] = val
        if pct:
            ws[coord].number_format = "0%"

    # ── 문서 메타 ──
    year, week = info["year"], info["week"]
    doc_no = inp.get("doc_no") or f"BE-{year}-W{week:02d}-{inp.get('doc_no_suffix', '01')}"
    S(w1, "C4", doc_no); S(w1, "E4", info["week_str"]); S(w1, "G4", info["period"])
    S(w1, "I4", info["report_due"]); S(w1, "K4", info["ack_due"])
    S(w1, "C5", inp["recipient"]); S(w1, "E5", inp.get("dept_role", ""))
    S(w1, "G5", inp.get("sender", "대표 서창환"))
    S(w1, "I5", inp.get("reviewer", "팀장 서창환"))
    S(w1, "K5", inp.get("doc_grade", "사내"))

    # ── §1 ──
    S(w1, "C8", inp.get("background", ""))
    S(w1, "C9", inp.get("goal", ""))
    S(w1, "C10", inp.get("loss", ""))

    # ── §2 ── (포함범위 C14..C21, 최대 8개)
    for i, t in enumerate(inp.get("scope_include", [])[:8]):
        S(w1, f"C{14 + i}", t)
    S(w1, "C22", inp.get("scope_exclude", ""))
    S(w1, "C23", inp.get("precondition", ""))
    S(w1, "C24", inp.get("scope_change", ""))

    # ── §3 지시업무 ──
    wsum = 0.0
    for i, t in enumerate(tasks):
        r = 28 + i
        S(w1, f"C{r}", t["name"])
        S(w1, f"D{r}", t.get("purpose", ""))
        S(w1, f"E{r}", t.get("done", ""))
        S(w1, f"F{r}", t.get("deliverable", ""))
        S(w1, f"G{r}", t.get("location", "브링 구글 드라이브"))
        deadline = t.get("deadline") or (info["friday"] if inp.get("unified_friday_deadline", True) else "")
        S(w1, f"H{r}", deadline)
        S(w1, f"I{r}", t.get("priority", "보통"))
        S(w1, f"J{r}", t.get("hours", ""))
        w = t.get("weight", 0)
        S(w1, f"K{r}", w, pct=True)
        wsum += w

    # 합계 / 발행검증 수식
    S(w1, f"J{sum1}", f"=SUM(J28:J{last})")
    S(w1, f"K{sum1}", f"=SUM(K28:K{last})")
    S(w1, f"C{issue1}",
      f'=IF(COUNTA(C28:C{last})=0,"업무 미입력",'
      f'IF(COUNTA(E28:E{last})<COUNTA(C28:C{last}),"발행 불가 — 완료기준 누락 업무 있음",'
      f'IF(COUNTA(F28:F{last})<COUNTA(C28:C{last}),"발행 불가 — 산출물 누락 업무 있음",'
      f'IF(SUM(K28:K{last})<>1,"발행 불가 — 가중치 합계 "&TEXT(SUM(K28:K{last}),"0%")&" (100% 필요)",'
      f'"발행 가능 — 완료기준·산출물·가중치 충족"))))')

    # 우선순위 드롭다운(최우선 포함)
    w1.data_validations.dataValidation = [
        dv for dv in w1.data_validations.dataValidation
        if not str(dv.sqref).startswith("I28")
    ]
    dv = DataValidation(type="list", formula1='"최우선,높음,보통,낮음"', allow_blank=True)
    dv.sqref = f"I28:I{last}"
    w1.add_data_validation(dv)

    # ── 결재란 ──
    ap = inp.get("approvers", {})
    S(w1, f"C{sign1}", ap.get("author", "서 창 환"))
    S(w1, f"E{sign1}", ap.get("reviewer", "서 창 환"))
    S(w1, f"G{sign1}", ap.get("approver", "서 창 환"))
    S(w1, f"I{sign1}", inp["recipient"])
    S(w1, f"K{sign1}", info["approve_date"])

    # ── 02_주간완료보고서: 업무 연동 & 합계 ──
    for i in range(N):
        r = 12 + i
        sr = 28 + i
        S(w2, f"D{r}", f"=IF('01_업무지시서'!C{sr}=\"\",\"\",'01_업무지시서'!C{sr})")
        S(w2, f"E{r}", f"=IF('01_업무지시서'!E{sr}=\"\",\"\",'01_업무지시서'!E{sr})")
        S(w2, f"F{r}", f"=IF('01_업무지시서'!K{sr}=\"\",0,'01_업무지시서'!K{sr})")
    S(w2, f"F{sum2}", f"=SUM(F12:F{lr})")
    S(w2, f"G{sum2}", f"=SUMPRODUCT(F12:F{lr},H12:H{lr})")
    S(w2, f"H{sum2}", f'=COUNTIF(G12:G{lr},"완료")')
    S(w2, f"I{sum2}", f"=SUM(I12:I{lr})")
    S(w2, f"J{sum2}",
      f'="완료 "&COUNTIF(G12:G{lr},"완료")&"건 / 진행중 "&COUNTIF(G12:G{lr},"진행중")'
      f'&"건 / 미착수 "&COUNTIF(G12:G{lr},"미착수")&"건"')
    S(w2, "D5", f"=SUMPRODUCT(F12:F{lr},H12:H{lr})")
    S(w2, "J5", f"=SUM(I12:I{lr})")
    S(w2, "H5", f"=SUM('01_업무지시서'!J28:J{last})")

    # 02 데이터검증 재설정 (M 만큼 이동)
    w2.data_validations.dataValidation = []
    for f1, sq in [
        ('"충족,미충족"', f"I{21 + M}:I{27 + M}"),
        ('"합격,보완,불합격"', f"L{21 + M}:L{27 + M}"),
        ('"승인,보류,반려"', f"K{37 + M}:K{40 + M} L{31 + M}:L{33 + M}"),
        ('"합격,조건부합격,보완요청,불합격"', f"K{47 + M}"),
        ('"완료,진행중,미착수,보류,취소"', f"G12:G{lr}"),
    ]:
        d = DataValidation(type="list", formula1=f1, allow_blank=True)
        d.sqref = sq
        w2.add_data_validation(d)

    # 가중치 검증
    if abs(wsum - 1.0) > 0.001:
        print(f"[경고] 가중치 합계가 {wsum:.3f} 입니다 (100% 필요). 발행 전 조정하세요.", file=sys.stderr)

    if not out_path:
        rc = inp["recipient"].replace(" ", "")
        out_path = f"BRING_주간_{rc}_{info['week_str']}.xlsx"
    wb.save(out_path)
    return out_path, info, wsum


def main():
    ap = argparse.ArgumentParser(description="BRING 주간 업무지시서 생성기")
    ap.add_argument("input", help="지시서 입력 JSON 경로")
    ap.add_argument("-o", "--out", help="출력 xlsx 경로 (기본: BRING_주간_<이름>_<주차>.xlsx)")
    ap.add_argument("--template", help=f"빈 템플릿 경로 (기본: {DEFAULT_TEMPLATE})")
    args = ap.parse_args()
    with open(args.input, encoding="utf-8") as fp:
        inp = json.load(fp)
    out, info, wsum = build(inp, args.out, args.template)
    print(f"생성 완료: {out}")
    print(f"  주차: {info['week_str']} ({info['period']}) / 보고 {info['report_due']}")
    print(f"  업무 {len(inp['tasks'])}개 / 가중치 합계 {round(wsum * 100)}%")


if __name__ == "__main__":
    main()
